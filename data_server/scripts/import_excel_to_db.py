"""
Import sites from Excel file to PostgreSQL database
Enhanced version with preview, merge/replace modes, and detailed logging
"""
import asyncio
import sys
import argparse
from pathlib import Path
from datetime import datetime
from typing import List, Tuple

# Add parent directory to path to import app modules
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from pyproj import Transformer
from sqlalchemy import select
from app.database import AsyncSessionLocal
from app.models.permanent_site import PermanentSite
from app.config import get_settings


class ImportLogger:
    """Logger that writes to both console and file"""
    def __init__(self, log_file: str = None):
        self.log_file = log_file
        self.logs = []
        
    def log(self, message: str):
        """Log message to console and buffer"""
        print(message)
        self.logs.append(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | {message}")
        
    def save(self):
        """Save logs to file if specified"""
        if self.log_file:
            with open(self.log_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(self.logs))
            print(f"\n📝 Log saved to: {self.log_file}")


def to_number(v):
    """Convert value to number, return None if invalid"""
    try:
        if isinstance(v, (int, float)):
            return float(v)
        if v is None or str(v).strip() == "":
            return None
        n = float(str(v).strip())
        return n if not (n != n) else None  # Check for NaN
    except (ValueError, TypeError):
        return None


def clean_str(v):
    """Clean string value"""
    if v is None:
        return ""
    s = str(v).strip()
    return s if s else ""


def pick(row, keys):
    """Return first non-empty value from list of possible column names"""
    for k in keys:
        if k in row and row[k]:
            s = clean_str(row[k])
            if s:
                return s
    return ""


def transform_coordinates(x, y):
    """Transform ITM (EPSG:2039) coordinates to WGS84 (lat, lng)"""
    if x is None or y is None:
        return None, None
    
    try:
        # Use pyproj to transform from ITM to WGS84
        transformer = Transformer.from_crs("EPSG:2039", "EPSG:4326", always_xy=True)
        lng, lat = transformer.transform(x, y)
        return lat, lng
    except Exception as e:
        return None, None


def parse_excel(excel_path: str, logger: ImportLogger) -> Tuple[List[PermanentSite], List[str]]:
    """Parse Excel file and return list of sites to import + errors"""
    
    logger.log(f"📖 Reading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    logger.log(f"📋 Found {len(headers)} columns")
    
    # Read all rows
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = {}
        for idx, value in enumerate(row):
            if idx < len(headers):
                row_dict[headers[idx]] = value
        rows.append((row_idx, row_dict))
    
    logger.log(f"📊 Found {len(rows)} rows in Excel")
    
    # Transform to PermanentSite objects
    sites_to_create = []
    errors = []
    
    for row_idx, r in rows:
        # Get coordinates (support multiple column name formats)
        # Note: In ITM (EPSG:2039), X is Easting (east-west) and Y is Northing (north-south)
        # The Excel columns might be labeled confusingly as "קו אורך X" and "קו רוחב Y"
        # We need to identify which values are Easting (smaller, ~180k-200k) vs Northing (larger, ~500k-800k)
        
        coord1 = to_number(pick(r, ["קו אורך", "קו אורך X", "קו אורך (Y)", "קו-אורך", "אורך", "Y"]))
        coord2 = to_number(pick(r, ["קו רוחב", "קו רוחב Y", "קו רוחב (X)", "קו-רוחב", "רוחב", "X"]))
        
        if coord1 is None or coord2 is None:
            error_msg = f"⚠️  Row {row_idx}: Missing coordinates"
            logger.log(error_msg)
            errors.append(error_msg)
            continue
        
        # Determine which is Easting (X) and which is Northing (Y) based on typical ITM ranges
        # Easting: typically 100,000 - 300,000
        # Northing: typically 500,000 - 800,000
        if coord1 > coord2:
            # coord1 is larger, likely Northing (Y)
            x = coord2  # Easting
            y = coord1  # Northing
        else:
            # coord2 is larger, likely Northing (Y)
            x = coord1  # Easting
            y = coord2  # Northing
        
        # Transform coordinates
        lat, lng = transform_coordinates(x, y)
        
        if lat is None or lng is None:
            error_msg = f"⚠️  Row {row_idx}: Invalid coordinates ({x}, {y})"
            logger.log(error_msg)
            errors.append(error_msg)
            continue
        
        # Get other fields
        name = pick(r, ["אתר", "שם אתר", "שם"])
        if not name:
            error_msg = f"⚠️  Row {row_idx}: Missing name"
            logger.log(error_msg)
            errors.append(error_msg)
            continue
        
        category = pick(r, ["קטגוריה", "קטגוריה ראשית", "קטגוריה ראשית "])
        sub_category = pick(r, ["תת קטגוריה", "תת-קטגוריה", "תת קטגוריה ", "תתקטגוריה"])
        site_type = pick(r, ["סוג אתר", "סוג"])
        district = pick(r, ["רובע", "רובע ", "רובע  ", "אזור", "שכונה"])
        street = pick(r, ["רחוב", "שם רחוב"])
        house_number = pick(r, ["מספר בית", "מספר", "בית"])
        phone = pick(r, ["טלפון איש קשר", "טלפון ", "מספר טלפון", "נייד", "פלאפון"])
        contact_name = pick(r, ["איש קשר", "אחראי", "אחראי/ת", "שם איש קשר", "איש קשר / אחראי"])
        description = pick(r, ["תאור כללי", "תיאור כללי", "תיאור", "הערות", "מידע נוסף"])
        
        site = PermanentSite(
            name=name,
            category=category or None,
            sub_category=sub_category or None,
            type=site_type or None,
            district=district or None,
            street=street or None,
            house_number=house_number or None,
            contact_name=contact_name or None,
            phone=phone or None,
            description=description or None,
            lat=lat,
            lng=lng
        )
        
        sites_to_create.append(site)
    
    logger.log(f"✅ Prepared {len(sites_to_create)} valid sites")
    logger.log(f"⚠️  Skipped {len(errors)} rows with errors")
    
    return sites_to_create, errors


async def preview_import(excel_path: str, logger: ImportLogger):
    """Preview what will be imported without making changes"""
    
    logger.log("\n" + "="*60)
    logger.log("🔍 PREVIEW MODE - No changes will be made")
    logger.log("="*60 + "\n")
    
    # Parse Excel
    sites_to_create, errors = parse_excel(excel_path, logger)
    
    # Check existing database
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(PermanentSite))
        existing_sites = result.scalars().all()
        existing_names = {site.name for site in existing_sites}
        
        # Check for duplicates
        new_sites = []
        duplicate_sites = []
        
        for site in sites_to_create:
            if site.name in existing_names:
                duplicate_sites.append(site.name)
            else:
                new_sites.append(site.name)
        
        logger.log("\n📊 IMPORT PREVIEW:")
        logger.log(f"  • Current DB sites: {len(existing_sites)}")
        logger.log(f"  • Sites in Excel: {len(sites_to_create)}")
        logger.log(f"  • New sites to add: {len(new_sites)}")
        logger.log(f"  • Duplicate names found: {len(duplicate_sites)}")
        logger.log(f"  • Errors: {len(errors)}")
        
        if duplicate_sites:
            logger.log("\n⚠️  DUPLICATE NAMES (first 10):")
            for name in duplicate_sites[:10]:
                logger.log(f"     - {name}")
        
        if new_sites:
            logger.log("\n✅ NEW SITES (first 10):")
            for name in new_sites[:10]:
                logger.log(f"     - {name}")
        
        logger.log("\n💡 To execute import, run without --preview flag")


async def import_sites_merge(excel_path: str, logger: ImportLogger):
    """Merge mode: Add new sites, skip duplicates"""
    
    logger.log("\n" + "="*60)
    logger.log("🔀 MERGE MODE - Adding new sites, keeping existing")
    logger.log("="*60 + "\n")
    
    # Parse Excel
    sites_to_create, errors = parse_excel(excel_path, logger)
    
    # Import
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(PermanentSite))
        existing_sites = result.scalars().all()
        existing_names = {site.name for site in existing_sites}
        
        # Filter out duplicates
        new_sites = [site for site in sites_to_create if site.name not in existing_names]
        skipped_duplicates = len(sites_to_create) - len(new_sites)
        
        if new_sites:
            session.add_all(new_sites)
            await session.commit()
            logger.log(f"✅ Added {len(new_sites)} new sites")
        else:
            logger.log("ℹ️  No new sites to add")
        
        if skipped_duplicates > 0:
            logger.log(f"⏭️  Skipped {skipped_duplicates} duplicate sites")
        
        logger.log(f"\n📊 FINAL STATUS:")
        logger.log(f"  • Total sites in DB: {len(existing_sites) + len(new_sites)}")


async def import_sites_replace(excel_path: str, logger: ImportLogger):
    """Replace mode: Delete all existing sites and import fresh"""
    
    logger.log("\n" + "="*60)
    logger.log("🔄 REPLACE MODE - Deleting existing and importing fresh")
    logger.log("="*60 + "\n")
    
    # Parse Excel
    sites_to_create, errors = parse_excel(excel_path, logger)
    
    # Import
    async with AsyncSessionLocal() as session:
        # Delete existing
        result = await session.execute(select(PermanentSite))
        existing_sites = result.scalars().all()
        
        if existing_sites:
            logger.log(f"🗑️  Deleting {len(existing_sites)} existing sites...")
            for site in existing_sites:
                await session.delete(site)
            await session.commit()
            logger.log("✅ Deleted existing sites")
        
        # Add new
        if sites_to_create:
            session.add_all(sites_to_create)
            await session.commit()
            logger.log(f"✅ Imported {len(sites_to_create)} new sites")
        
        logger.log(f"\n📊 FINAL STATUS:")
        logger.log(f"  • Total sites in DB: {len(sites_to_create)}")


async def main():
    parser = argparse.ArgumentParser(
        description="Import sites from Excel to PostgreSQL",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Preview without changes
  python import_excel_to_db.py --preview
  
  # Merge mode (add new, keep existing)
  python import_excel_to_db.py --mode merge
  
  # Replace mode (delete all, import fresh)
  python import_excel_to_db.py --mode replace
  
  # With custom Excel file
  python import_excel_to_db.py --file /path/to/Omer_GIS_Reorganized_Final.xlsx
  
  # With log file
  python import_excel_to_db.py --log-file import_2026-01-11.log
        """
    )
    
    parser.add_argument(
        '--preview',
        action='store_true',
        help='Preview changes without executing (default: False)'
    )
    
    parser.add_argument(
        '--mode',
        choices=['merge', 'replace'],
        default='merge',
        help='Import mode: merge (add new) or replace (delete all + import) (default: merge)'
    )
    
    parser.add_argument(
        '--file',
        type=str,
        default='/app/Omer_GIS_Reorganized_Final.xlsx',
        help='Path to Excel file (default: /app/Omer_GIS_Reorganized_Final.xlsx)'
    )
    
    parser.add_argument(
        '--log-file',
        type=str,
        default=None,
        help='Save log to file (default: console only)'
    )
    
    args = parser.parse_args()
    
    # Setup logger
    if args.log_file is None and not args.preview:
        # Auto-generate log file name if not preview
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        args.log_file = f"/app/import_{timestamp}.log"
    
    logger = ImportLogger(args.log_file)
    
    # Check file exists
    excel_path = Path(args.file)
    if not excel_path.exists():
        logger.log(f"❌ Excel file not found: {excel_path}")
        logger.log("Please copy Omer_GIS_Reorganized_Final.xlsx to the container first:")
        logger.log("docker cp front/public/Omer_GIS_Reorganized_Final.xlsx omeropsmap_data_server:/app/Omer_GIS_Reorganized_Final.xlsx")
        return
    
    # Execute based on mode
    try:
        if args.preview:
            await preview_import(str(excel_path), logger)
        elif args.mode == 'merge':
            await import_sites_merge(str(excel_path), logger)
        elif args.mode == 'replace':
            await import_sites_replace(str(excel_path), logger)
        
        logger.log("\n✅ Operation completed successfully!")
        
    except Exception as e:
        logger.log(f"\n❌ Error: {e}")
        import traceback
        logger.log(traceback.format_exc())
    finally:
        logger.save()


if __name__ == "__main__":
    asyncio.run(main())
