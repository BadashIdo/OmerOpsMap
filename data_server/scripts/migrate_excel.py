"""
Migration script to import data from Excel file to PostgreSQL database
"""
import asyncio
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from pyproj import Transformer
from sqlalchemy import text

from app.database import AsyncSessionLocal, engine, Base
from app.models import PermanentSite


# EPSG:2039 (Israel TM Grid / ITM) to WGS84
transformer = Transformer.from_crs("EPSG:2039", "EPSG:4326", always_xy=True)


def to_number(v):
    """Convert value to number"""
    try:
        if isinstance(v, (int, float)):
            return float(v)
        n = float(str(v).strip())
        return n if not (n != n) else None  # Check for NaN
    except (ValueError, TypeError):
        return None


def clean_str(v):
    """Clean string value"""
    s = str(v or "").strip()
    return s if s else ""


def pick(row_dict, keys):
    """Get first non-empty value from list of possible keys"""
    for k in keys:
        v = row_dict.get(k)
        s = clean_str(v)
        if s:
            return s
    return ""


async def migrate_excel_to_db(excel_path: str):
    """Migrate Excel data to PostgreSQL database"""
    
    print(f"Loading Excel file: {excel_path}")
    
    # Load Excel file
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Convert rows to dictionaries
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            rows.append(row_dict)
        
        print(f"Found {len(rows)} rows in Excel")
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return
    
    # Create database tables
    print("Creating database tables...")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    
    # Process and insert sites
    sites_created = 0
    sites_skipped = 0
    
    async with AsyncSessionLocal() as session:
        for idx, row in enumerate(rows, start=1):
            try:
                # Get coordinates (ITM format)
                y = to_number(pick(row, ["קו אורך", "קו-אורך", "אורך", "Y"]))
                x = to_number(pick(row, ["קו רוחב", "קו-רוחב", "רוחב", "X"]))
                
                if x is None or y is None:
                    print(f"Row {idx}: Skipping - missing coordinates")
                    sites_skipped += 1
                    continue
                
                # Convert ITM to WGS84
                lng, lat = transformer.transform(x, y)
                
                # Extract fields
                name = pick(row, ["אתר", "שם אתר", "שם"])
                if not name:
                    print(f"Row {idx}: Skipping - missing name")
                    sites_skipped += 1
                    continue
                
                category = pick(row, ["קטגוריה", "קטגוריה ראשית"])
                sub_category = pick(row, ["תת קטגוריה", "תת-קטגוריה", "תת קטגוריה ", "תתקטגוריה"])
                site_type = pick(row, ["סוג אתר", "סוג"])
                district = pick(row, ["רובע", "רובע ", "רובע  ", "אזור", "שכונה"])
                street = pick(row, ["רחוב", "שם רחוב"])
                house_number = pick(row, ["מספר בית", "מספר", "בית"])
                phone = pick(row, ["טלפון איש קשר", "טלפון ", "מספר טלפון", "נייד", "פלאפון"])
                contact_name = pick(row, ["איש קשר", "אחראי", "אחראי/ת", "שם איש קשר", "איש קשר / אחראי"])
                description = pick(row, ["תאור כללי", "תיאור כללי", "תיאור", "הערות", "מידע נוסף"])
                
                # Create site
                site = PermanentSite(
                    name=name,
                    category=category or None,
                    sub_category=sub_category or None,
                    type=site_type or None,
                    district=district or None,
                    street=street or None,
                    house_number=house_number or None,
                    contact_name=contact_name or None,
                    phone=phone or None,
                    description=description or None,
                    lat=lat,
                    lng=lng
                )
                
                session.add(site)
                sites_created += 1
                
                if sites_created % 10 == 0:
                    print(f"Processed {sites_created} sites...")
                
            except Exception as e:
                print(f"Row {idx}: Error - {e}")
                sites_skipped += 1
                continue
        
        # Commit all sites
        await session.commit()
    
    print(f"\n=== Migration Complete ===")
    print(f"Sites created: {sites_created}")
    print(f"Sites skipped: {sites_skipped}")
    print(f"Total rows: {len(rows)}")


async def main():
    # Default Excel path (relative to project root)
    default_path = Path(__file__).parent.parent.parent / "front" / "public" / "sites.xlsx"
    
    excel_path = sys.argv[1] if len(sys.argv) > 1 else str(default_path)
    
    if not Path(excel_path).exists():
        print(f"Error: Excel file not found at {excel_path}")
        print("Usage: python migrate_excel.py [path/to/sites.xlsx]")
        sys.exit(1)
    
    await migrate_excel_to_db(excel_path)


if __name__ == "__main__":
    asyncio.run(main())

