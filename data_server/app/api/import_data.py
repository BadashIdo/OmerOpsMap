"""
API endpoints for data import from Excel files
"""
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, status, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import List, Literal, Tuple
import tempfile
import shutil
from pathlib import Path
from datetime import datetime, timezone
import math
import openpyxl

from app.database import get_db
from app.auth.jwt import get_current_admin
from app.models.admin import Admin
from app.models.permanent_site import PermanentSite
from app.models.temporary_site import TemporarySite
from pydantic import BaseModel

import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "scripts"))
from import_excel_to_db import parse_excel, ImportLogger, clean_str, to_number


def haversine_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate distance in meters between two coordinates"""
    if lat1 is None or lon1 is None or lat2 is None or lon2 is None:
        return float('inf')
    R = 6371000  # Radius of earth in meters
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    
    a = math.sin(delta_phi / 2.0) ** 2 + \
        math.cos(phi1) * math.cos(phi2) * \
        math.sin(delta_lambda / 2.0) ** 2
    
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c


router = APIRouter(prefix="/api/admin/import", tags=["import"])


# ── Pydantic models ───────────────────────────────────────────────────────────

class ImportPreviewResponse(BaseModel):
    total_rows: int
    valid_sites: int
    errors: List[str]
    new_sites: List[str]
    duplicate_sites: List[str]
    current_db_count: int

    class Config:
        from_attributes = True


class ImportExecuteResponse(BaseModel):
    success: bool
    message: str
    sites_added: int
    sites_deleted: int
    sites_skipped: int
    total_in_db: int

    class Config:
        from_attributes = True


# ── Sheet-type detection ──────────────────────────────────────────────────────

TEMP_MARKER_COLS = {"דחיפות", "סטטוס", "תאריך התחלה", "תאריך סיום"}
PERM_MARKER_COLS = {"תת-קטגוריה", "שכונה/רובע", "רחוב", "מספר בית"}


def _detect_sheet_type(wb) -> str:
    """Return 'permanent', 'temporary', or 'unknown'."""
    sheet_names = [ws.title for ws in wb.worksheets]
    if "אירועים זמניים" in sheet_names:
        return "temporary"
    if "אתרים קבועים" in sheet_names:
        return "permanent"

    # Fall back to column-header sniffing on the active sheet
    ws = wb.active
    headers = {cell.value for cell in ws[1] if cell.value}
    if TEMP_MARKER_COLS & headers:
        return "temporary"
    if PERM_MARKER_COLS & headers:
        return "permanent"
    return "unknown"


# ── Temporary-site parser ─────────────────────────────────────────────────────

_PRIORITY_MAP = {
    "נמוכה": "low", "בינונית": "medium", "גבוהה": "high", "קריטית": "critical",
    "low": "low", "medium": "medium", "high": "high", "critical": "critical",
}
_STATUS_MAP = {
    "פעיל": "active", "מושהה": "paused", "הסתיים": "resolved",
    "מתוכנן": "active", "פג תוקף": "resolved", "בוטל": "resolved",
    "active": "active", "paused": "paused", "resolved": "resolved",
}


def _parse_dt(val):
    if val is None:
        return None
    if hasattr(val, "year"):
        return val
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(val).strip(), fmt)
        except ValueError:
            pass
    return None


def parse_temporary_excel(excel_path: str, logger: ImportLogger) -> Tuple[List[TemporarySite], List[str]]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    ws = next((s for s in wb.worksheets if s.title == "אירועים זמניים"), wb.active)
    headers = [cell.value for cell in ws[1]]

    sites: List[TemporarySite] = []
    errors: List[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        r = {headers[i]: v for i, v in enumerate(row) if i < len(headers) and headers[i]}
        if not any(v for v in r.values() if v is not None):
            continue

        name = clean_str(r.get("שם", ""))
        if not name:
            errors.append(f"⚠️ שורה {row_idx}: חסר שם")
            continue

        lat = to_number(r.get("קו רוחב"))
        lng = to_number(r.get("קו אורך"))
        if lat is None or lng is None:
            errors.append(f"⚠️ שורה {row_idx}: חסרות קואורדינטות")
            continue

        end_date = _parse_dt(r.get("תאריך סיום"))
        if end_date is None:
            errors.append(f"⚠️ שורה {row_idx}: חסר תאריך סיום")
            continue

        start_date = _parse_dt(r.get("תאריך התחלה")) or datetime.now(timezone.utc)

        priority_raw = clean_str(r.get("דחיפות", ""))
        status_raw   = clean_str(r.get("סטטוס", ""))

        category = clean_str(r.get("קטגוריה", "")) or None
        if not category:
            errors.append(f"⚠️ שורה {row_idx}: חסרה קטגוריה")
            continue

        sub_category = clean_str(r.get("תת קטגוריה", r.get("תת-קטגוריה", ""))) or None
        if not sub_category:
            errors.append(f"⚠️ שורה {row_idx}: חסרה תת קטגוריה")
            continue

        site_type = clean_str(r.get("סוג אתר", r.get("סוג", ""))) or None
        
        district = clean_str(r.get("רובע", r.get("אזור", r.get("שכונה", "")))) or None
        if not district:
            errors.append(f"⚠️ שורה {row_idx}: חסר רובע")
            continue

        street = clean_str(r.get("רחוב", r.get("שם רחוב", ""))) or None
        house_number = clean_str(r.get("מספר בית", r.get("מספר", ""))) or None

        sites.append(TemporarySite(
            name=name,
            category=category,
            sub_category=sub_category,
            type=site_type,
            district=district,
            street=street,
            house_number=house_number,
            description=clean_str(r.get("תיאור", "")) or None,
            lat=lat,
            lng=lng,
            start_date=start_date,
            end_date=end_date,
            status=_STATUS_MAP.get(status_raw, "active"),
            contact_name=clean_str(r.get("שם איש קשר", "")) or None,
            phone=clean_str(r.get("טלפון", "")) or None,
        ))

    return sites, errors


# ── Helpers ───────────────────────────────────────────────────────────────────

def _validate_type_match(detected: str, selected: str):
    if detected == "unknown":
        return  # Can't tell — let it through
    if detected != selected:
        type_he = {"permanent": "אתרים קבועים", "temporary": "אירועים זמניים"}
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"הקובץ שהועלה מכיל {type_he.get(detected, detected)} "
                f"אך נבחר יבוא של {type_he.get(selected, selected)}. "
                "נא לבחור את הסוג הנכון."
            ),
        )


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/preview", response_model=ImportPreviewResponse)
async def preview_import(
    file: UploadFile = File(...),
    site_type: Literal["permanent", "temporary"] = Query(default="permanent"),
    db: AsyncSession = Depends(get_db),
    admin: Admin = Depends(get_current_admin),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="נא להעלות קובץ Excel (.xlsx או .xls)")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        _validate_type_match(_detect_sheet_type(wb), site_type)

        logger = ImportLogger(log_file=None)

        if site_type == "permanent":
            sites_to_create, errors = parse_excel(tmp_path, logger)
            result = await db.execute(select(PermanentSite))
            existing_sites = result.scalars().all()
            db_count = len(existing_sites)
        else:
            sites_to_create, errors = parse_temporary_excel(tmp_path, logger)
            result = await db.execute(select(TemporarySite))
            existing_sites = result.scalars().all()
            db_count = len(existing_sites)

        from collections import defaultdict
        existing_by_name = defaultdict(list)
        for s in existing_sites:
            existing_by_name[s.name].append(s)

        new_sites = []
        dup_sites = []
        
        for s in sites_to_create:
            is_duplicate = False
            if s.name in existing_by_name:
                for existing_s in existing_by_name[s.name]:
                    dist = haversine_distance(s.lat, s.lng, existing_s.lat, existing_s.lng)
                    if dist <= 5.0:
                        is_duplicate = True
                        break
            
            if is_duplicate:
                dup_sites.append(s.name)
            else:
                new_sites.append(s.name)
                # To prevent duplicates within the file itself
                existing_by_name[s.name].append(s)

        return ImportPreviewResponse(
            total_rows=len(sites_to_create) + len(errors),
            valid_sites=len(sites_to_create),
            errors=errors,
            new_sites=new_sites[:50],
            duplicate_sites=dup_sites[:50],
            current_db_count=db_count,
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"שגיאה בניתוח הקובץ: {str(e)}")
    finally:
        Path(tmp_path).unlink(missing_ok=True)


@router.post("/execute", response_model=ImportExecuteResponse)
async def execute_import(
    file: UploadFile = File(...),
    site_type: Literal["permanent", "temporary"] = Query(default="permanent"),
    db: AsyncSession = Depends(get_db),
    admin: Admin = Depends(get_current_admin),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="נא להעלות קובץ Excel (.xlsx או .xls)")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        shutil.copyfileobj(file.file, tmp)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        _validate_type_match(_detect_sheet_type(wb), site_type)

        logger = ImportLogger(log_file=None)

        if site_type == "permanent":
            sites_to_create, _ = parse_excel(tmp_path, logger)
            Model = PermanentSite
        else:
            sites_to_create, _ = parse_temporary_excel(tmp_path, logger)
            Model = TemporarySite

        result = await db.execute(select(Model))
        existing = result.scalars().all()

        sites_added = sites_deleted = sites_skipped = 0

        from collections import defaultdict
        existing_by_name = defaultdict(list)
        for s in existing:
            existing_by_name[s.name].append(s)

        new_sites = []
        for s in sites_to_create:
            is_duplicate = False
            if s.name in existing_by_name:
                for existing_s in existing_by_name[s.name]:
                    dist = haversine_distance(s.lat, s.lng, existing_s.lat, existing_s.lng)
                    if dist <= 5.0:
                        is_duplicate = True
                        break
            
            if not is_duplicate:
                new_sites.append(s)
                existing_by_name[s.name].append(s)

        sites_skipped = len(sites_to_create) - len(new_sites)
        if new_sites:
            db.add_all(new_sites)
            sites_added = len(new_sites)

        await db.commit()

        result = await db.execute(select(Model))
        total = len(result.scalars().all())

        return ImportExecuteResponse(
            success=True,
            message="יבוא הושלם בהצלחה",
            sites_added=sites_added,
            sites_deleted=sites_deleted,
            sites_skipped=sites_skipped,
            total_in_db=total,
        )

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"שגיאה ביבוא הנתונים: {str(e)}")
    finally:
        Path(tmp_path).unlink(missing_ok=True)
