from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from datetime import datetime, timezone
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import get_db
from app.auth.jwt import get_current_admin
from app.models.admin import Admin
from app.models.permanent_site import PermanentSite
from app.models.temporary_site import TemporarySite
from app.models.feedback import Feedback
from app.models.external_feature import ExternalFeature, IntegrationRun

router = APIRouter(prefix="/api/admin/export", tags=["export"])


def _style_header_row(ws, fill_color: str):
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _fmt(dt) -> str:
    if dt is None:
        return ""
    if hasattr(dt, "strftime"):
        return dt.strftime("%Y-%m-%d %H:%M")
    return str(dt)


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 100) # Cap at 100 to prevent giant JSON columns


async def _export_permanent_sites(db: AsyncSession, wb, is_first: bool = False):
    result = await db.execute(select(PermanentSite).order_by(PermanentSite.id))
    sites = result.scalars().all()
    ws = wb.active if is_first else wb.create_sheet("אתרים קבועים")
    if is_first:
        ws.title = "אתרים קבועים"
    ws.append([
        "ID", "שם", "קטגוריה", "תת-קטגוריה", "סוג", "שכונה/רובע",
        "רחוב", "מספר בית", "שם איש קשר", "טלפון", "תיאור",
        "קו רוחב", "קו אורך", "נוצר", "עודכן",
    ])
    _style_header_row(ws, "1565C0")
    for s in sites:
        ws.append([
            s.id, s.name, s.category or "", s.sub_category or "",
            s.type or "", s.district or "", s.street or "",
            s.house_number or "", s.contact_name or "", s.phone or "",
            s.description or "", s.lat, s.lng,
            _fmt(s.created_at), _fmt(s.updated_at),
        ])
    _auto_width(ws)


async def _export_temporary_sites(db: AsyncSession, wb, is_first: bool = False):
    result = await db.execute(select(TemporarySite).order_by(TemporarySite.id))
    sites = result.scalars().all()
    ws = wb.active if is_first else wb.create_sheet("אירועים זמניים")
    if is_first:
        ws.title = "אירועים זמניים"
    ws.append([
        "ID", "שם", "קטגוריה", "תיאור", "דחיפות", "סטטוס",
        "תאריך התחלה", "תאריך סיום", "שם איש קשר", "טלפון",
        "קו רוחב", "קו אורך", "נוצר", "עודכן",
    ])
    _style_header_row(ws, "E65100")
    for t in sites:
        ws.append([
            t.id, t.name, t.category or "", t.description or "",
            t.priority.value if t.priority else "",
            t.status.value if t.status else "",
            _fmt(t.start_date), _fmt(t.end_date),
            t.contact_name or "", t.phone or "",
            t.lat, t.lng,
            _fmt(t.created_at), _fmt(t.updated_at),
        ])
    _auto_width(ws)


async def _export_feedback(db: AsyncSession, wb, is_first: bool = False):
    result = await db.execute(select(Feedback).order_by(Feedback.id))
    feedbacks = result.scalars().all()
    ws = wb.active if is_first else wb.create_sheet("דיווחי משתמשים")
    if is_first:
        ws.title = "דיווחי משתמשים"
    ws.append([
        "ID", "שם שולח", "נושא", "טלפון לזיהוי/קשר", "תיאור דיווח",
        "סטטוס טיפול", "הערות מנהל", "קו רוחב", "קו אורך", "תאריך דיווח", "תאריך עדכון"
    ])
    _style_header_row(ws, "2E7D32")
    for f in feedbacks:
        ws.append([
            f.id, f.name, f.topic, f.contact or "", f.description or "",
            f.status, f.admin_notes or "", f.lat, f.lng,
            _fmt(f.created_at), _fmt(f.updated_at)
        ])
    _auto_width(ws)


async def _export_admins(db: AsyncSession, wb, is_first: bool = False):
    result = await db.execute(select(Admin).order_by(Admin.id))
    admins = result.scalars().all()
    ws = wb.active if is_first else wb.create_sheet("מנהלי מערכת")
    if is_first:
        ws.title = "מנהלי מערכת"
    ws.append([
        "ID", "שם משתמש", "שם תצוגה", "אימייל", "סטטוס פעיל", "הרשאות (Role)",
        "נוצר ב", "התחברות אחרונה"
    ])
    _style_header_row(ws, "D32F2F")
    for a in admins:
        ws.append([
            a.id, a.username, a.display_name or "", a.email or "",
            "פעיל" if a.is_active else "לא פעיל", a.role,
            _fmt(a.created_at), _fmt(a.last_login)
        ])
    _auto_width(ws)


async def _export_external_features(db: AsyncSession, wb, is_first: bool = False):
    result = await db.execute(select(ExternalFeature).order_by(ExternalFeature.id))
    features = result.scalars().all()
    ws = wb.active if is_first else wb.create_sheet("שכבות חיצוניות")
    if is_first:
        ws.title = "שכבות חיצוניות"
    ws.append([
        "ID", "מקור", "מזהה חיצוני", "סוג (Kind)", "שם", "תיאור", "חומרה",
        "קו רוחב", "קו אורך", "מידע נוסף (Payload)", "תאריך שליפה", "תאריך פג תוקף"
    ])
    _style_header_row(ws, "607D8B")
    for f in features:
        payload_str = json.dumps(f.payload, ensure_ascii=False) if f.payload else ""
        ws.append([
            f.id, f.source.value if f.source else "", f.external_id or "", f.kind, f.name,
            f.description or "", f.severity or "", f.lat, f.lng,
            payload_str, _fmt(f.fetched_at), _fmt(f.expires_at)
        ])
    _auto_width(ws)


async def _export_integration_runs(db: AsyncSession, wb, is_first: bool = False):
    result = await db.execute(select(IntegrationRun).order_by(IntegrationRun.started_at.desc()))
    runs = result.scalars().all()
    ws = wb.active if is_first else wb.create_sheet("היסטוריית סנכרונים")
    if is_first:
        ws.title = "היסטוריית סנכרונים"
    ws.append([
        "מזהה ריצה (UUID)", "מקור סנכרון", "התחיל", "הסתיים", "הצלחה",
        "נוספו", "עודכנו", "נמחקו", "שגיאה"
    ])
    _style_header_row(ws, "546E7A")
    for r in runs:
        ws.append([
            str(r.id), r.source.value if r.source else "",
            _fmt(r.started_at), _fmt(r.finished_at),
            "כן" if r.ok else "לא",
            r.added, r.updated, r.removed, r.error or ""
        ])
    _auto_width(ws)


@router.get("/database")
async def export_database(
    table: str = Query("all", description="Table name to export or 'all'"),
    db: AsyncSession = Depends(get_db),
    admin: Admin = Depends(get_current_admin),
):
    wb = openpyxl.Workbook()

    if table == "all":
        await _export_permanent_sites(db, wb, is_first=True)
        await _export_temporary_sites(db, wb, is_first=False)
        await _export_feedback(db, wb, is_first=False)
        await _export_admins(db, wb, is_first=False)
        await _export_external_features(db, wb, is_first=False)
        await _export_integration_runs(db, wb, is_first=False)
    elif table == "permanent_sites":
        await _export_permanent_sites(db, wb, is_first=True)
    elif table == "temporary_sites":
        await _export_temporary_sites(db, wb, is_first=True)
    elif table == "feedback":
        await _export_feedback(db, wb, is_first=True)
    elif table == "admins":
        await _export_admins(db, wb, is_first=True)
    elif table == "external_features":
        await _export_external_features(db, wb, is_first=True)
    elif table == "integration_runs":
        await _export_integration_runs(db, wb, is_first=True)
    else:
        # Fallback if unknown table
        await _export_permanent_sites(db, wb, is_first=True)

    # ── Stream back as .xlsx ──────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"omeropsmap_export_{table}_{timestamp}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
